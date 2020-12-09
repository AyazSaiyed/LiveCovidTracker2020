#Ayaz Saiyed , 2020
# https://live-covid-tracker.herokuapp.com/

from django.shortcuts import render
from django.http import HttpResponseRedirect,JsonResponse
import json
import requests
from django.views.generic import View
# from django.utils.timezone import datetime
from .models import StateData

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl import load_workbook

from covidGraph import CountryWise



def about(request):
    return render(request,'about.html')


# import datetime
# LastUpdatedTime = datetime.datetime.now()
# print(LastUpdatedTime.strftime("%c"))
# LastUpdatedTime1 = LastUpdatedTime.strftime("%c")
import requests
from bs4 import BeautifulSoup
from datetime import datetime

URL = 'https://www.worldometers.info/coronavirus/#countries'

headers = {
      "User-Agents": 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.76 Safari/537.36'}
def finalAnalysis(request):
    CountryWise.CountryWise_exe()
    import _datetime
    LastUpdatedTime = _datetime.datetime.now()
    LastUpdatedTime1 = LastUpdatedTime.strftime("%c")
    CountryNameArray = []
    NumberofCountry = []
    TotalCasesArray = []
    NewCasesArray = []
    TotalDeathCasesArray = []
    NewDeathsArray = []
    TotalRecoveredCasesArray = []
    TotalActiveCasesArray = []
    SeriousCasesArray = []
    TotalCases1MPOPArray = []
    TotalDeaths1MPOPArray = []
    TotalTestsArray = []
    Test1MPOPArray = []
    TotalPopulationArray = []
    ContinentArray = []
    extract_countries = []
    def extract_global(soup):
      counts = [n.text for n in soup.find_all('div', class_= 'maincounter-number')]
      return counts



    def stats_global(soup):
      counts = extract_global(soup)
      data = {}
      data['totalCases'] = counts[0]
      data['totalDeaths'] = counts[1]
      data['totalRecovered'] = counts[2]

      # print("GLOBAL STATISTICS")
      # print()

      # print("Total Cases: %s" %data['totalCases'])
      # print("Total Deaths: %s" %data['totalDeaths'])
      # print("Total Recovered: %s" %data['totalRecovered'])
      print('')


    def get_country(country, data):
      # print("data ---",data)

      print("__________________")
      # print(country)
      print("__________________")
      # print()



    

    def extract_countries(soup):
      totalcase1 = []
      totanewcases1 = []
      countries = []
      table = soup.find('table')
      tableBody = table.find('tbody')

      rows = tableBody.find_all('tr')

      for row in rows:
        cols = row.find_all('td')
        cols = [ele.text.strip() for ele in cols]
        print("cols",cols)
        

        for i in cols:
            if i=='':
                print(" empty ")
            print(" value of i ",i)
        

        if not cols:
            print(" NullIsFound ")
            TRecovered = "None"
        else:
            TRecovered = cols[5]


        countrynumber = cols[0]
        countryreportnames = cols[1]
        TCases = cols[2]
        NewCases = cols[3]
        TDeaths = cols[4]
        NewDeaths = cols[5]
        TRecovered = cols[5]
        ActiveCases = cols[6]
        Total1MPop = cols[8]
        Deaths1MPop = cols[9]
        TotalTests = cols[10]
        Test1MPop = cols[11]
        TPopulation = cols[12]
        Test1M = cols[13]
        TotalPopulation = cols[14]
        Continent = cols[15]

        # workbook = Workbook()
        # workbook = load_workbook(filename="CovidData.xlsx")
        # sheet = workbook.active
        countryNumber1 = countrynumber
        countryName1 = countryreportnames
        cases1 = TCases
        NewCases1 = NewCases
        TotalDeathCases1 = TDeaths
        NewdeathCases1 = TRecovered
        TotalRecovered1 = ActiveCases
        TotalActiveCases1 = Total1MPop
        SeriousCases1 = Deaths1MPop
        TotalCases1MPOP1 = TotalTests
        Deaths1MPOP1 = Test1MPop
        TotalTests1 = TPopulation
        Tests1MPOP1 = Test1M
        Population1 = TotalPopulation
        # rows = [[countryNumber1,countryName1,cases1,NewCases1,TotalDeathCases1,NewdeathCases1,TotalRecovered1,TotalActiveCases1,SeriousCases1,TotalCases1MPOP1,Deaths1MPOP1,TotalTests1,Tests1MPOP1,Population1]]
        # for row in rows:
        # 	sheet.append(row)


        # workbook.save(filename="CovidData.xlsx")
        # for col in sheet.columns:
        # 	max_length = 0
        # 	column = col[0].column

        # 	for cell in col:
        # 		try:
        # 			if len(str(cell.value)) > max_length:
        # 				max_length = len(cell.value)
        # 		except:
        # 			pass
        # 	adjusted_width = (max_length + 2) * 1.2
        # 	sheet.column_dimensions[column].width = adjusted_width
        
        # workbook.save(filename="CovidData.xlsx")
        # print(" Data Saved ")
        # print(" ----------- ----------- -------------")
        # print(" ----------- ----------- -------------")

        # print(" Total Cases ",TCases)
        # print(" New Cases ",NewCases)  
        # print(" Total Deaths ",TDeaths)
        # print(" New Deaths ",TRecovered)
        # print(" Total Recovered ",ActiveCases)
        # # print(" Active Cases ",ActiveCases)

        # print(" Active Cases ",Total1MPop)
        # print(" Serious Critical ",Deaths1MPop)
        # print(" Total Cases/1 M Pop ",TotalTests)
        # print(" Deaths / 1M Pop ",Test1MPop)
        # print(" Total Tests ",TPopulation)
        # print(" Test / 1M Pop ",Test1M)
        # print(" Total Population ",TotalPopulation)
        # print(" Continent ",Continent)



        CountryNameArray.append(countryreportnames)
        NumberofCountry.append(countrynumber)    
        TotalCasesArray.append(TCases)
        NewCasesArray.append(NewCases)
        TotalDeathCasesArray.append(TDeaths)
        NewDeathsArray.append(TRecovered)
        TotalRecoveredCasesArray.append(ActiveCases)
        TotalActiveCasesArray.append(Total1MPop)
        SeriousCasesArray.append(Deaths1MPop)
        TotalCases1MPOPArray.append(TotalTests)
        TotalDeaths1MPOPArray.append(Test1MPop)
        TotalTestsArray.append(TPopulation)
        Test1MPOPArray.append(Test1M)
        TotalPopulationArray.append(TotalPopulation)
        ContinentArray.append(Continent)

        countries.append([ele for ele in cols if ele])

        for country in countries[:100]:
            if country[0].isnumeric():
                if len(country) > 5:
                    country.pop(0)
                get_country(country[0], country)
            else:
                get_country(country[0], country)


    web = requests.get(URL, headers, timeout=3)
    if web.status_code != 200:
        return None
    soup = BeautifulSoup(web.content, "html.parser")

    d = datetime.now().strftime("%m/%d/%Y -- %I:%M %p")

    print("Updated at: %s" %d)

    stats_global(soup)
    extract_countries(soup)


    return render(request,'index.html',{'CountryNameArray':CountryNameArray, 'NumberofCountry':NumberofCountry,'TotalCasesArray':TotalCasesArray,'NewCasesArray':NewCasesArray,'TotalDeathCasesArray':TotalDeathCasesArray,'NewDeathsArray':NewDeathsArray,'TotalRecoveredCasesArray':TotalRecoveredCasesArray,'TotalActiveCasesArray':TotalActiveCasesArray,'SeriousCasesArray':SeriousCasesArray,'TotalCases1MPOPArray':TotalCases1MPOPArray,'TotalDeaths1MPOPArray':TotalDeaths1MPOPArray,'TotalTestsArray':TotalTestsArray,'Test1MPOPArray':Test1MPOPArray,'TotalPopulationArray':TotalPopulationArray,'ContinentArray':ContinentArray,'UpdatedTime':LastUpdatedTime1})


    scrapper_data_covid()





