from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl import load_workbook

workbook = Workbook()


workbook = load_workbook(filename="CovidData.xlsx")
sheet = workbook.active

# Let's create some sample sales data
rows = [
			["Index","Country","Cases","New Cases","Total Death Cases"," New death Cases"," Total Recovered"," Total Active Cases"," Serious Cases "," Total Cases/1M POP","Deaths / 1M POP","Total Tests","Tests/1M POP", "Population"]
       ]


for row in rows:
	sheet.append(row)

workbook.save(filename="CovidData.xlsx")

workbook.close()


# ----




# import xlsxwriter 
  
# # Workbook() takes one, non-optional, argument  
# # which is the filename that we want to create. 
# workbook = xlsxwriter.Workbook('CovidData.xlsx') 
  
# # The workbook object is then used to add new  
# # worksheet via the add_worksheet() method. 
# worksheet = workbook.add_worksheet() 
  
# # Use the worksheet object to write 
# # # data via the write() method. 
# # worksheet.write('A1', 'Hello..') 
# # worksheet.write('B1', 'Geeks') 
# # worksheet.write('C1', 'For') 
# # worksheet.write('D1', 'Geeks') 
  
# # Finally, close the Excel file 
# # via the close() method. 
# workbook.close() 